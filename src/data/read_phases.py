"""
Map phase-labels.json videoIds to video names/metadata from the Excel database.

Searches all sheets in the Excel file for touchsurgery video UUIDs,
then joins with phase annotations.

Usage:
    python src/data/read_phases.py
    python src/data/read_phases.py --phases src/data/phase-labels.json --excel "own experiments/Cholec80/Non-Pituitary Operative Video Database - Foundational Model.xlsx"
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

EXCEL_PATH = _PROJECT_ROOT / "own experiments" / "Cholec80" / "Non-Pituitary Operative Video Database - Foundational Model.xlsx"
PHASES_PATH = _PROJECT_ROOT / "src" / "data" / "phase-labels.json"

_UUID_RE = re.compile(r"[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}", re.I)


def _extract_uuid(text: str) -> str | None:
    """Extract a UUID from a string (e.g. a touchsurgery URL)."""
    m = _UUID_RE.search(str(text))
    return m.group(0).lower() if m else None


def build_uuid_to_info(excel_path: Path = EXCEL_PATH) -> dict[str, dict]:
    """Scan all sheets in the Excel file and build a UUID → info mapping.

    For each row that contains a touchsurgery UUID, extracts the best
    available name and category/dataset info.

    Returns:
        {uuid: {"name": str, "dataset": str, "sheet": str}}
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(excel_path), read_only=True)
    uuid_map: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row (first row with string values)
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

        # Detect which columns have UUIDs/links and names
        # We'll scan every cell for UUIDs
        for row_idx, row in enumerate(rows[1:], start=1):
            cells = [str(c) if c else "" for c in row]
            row_text = " ".join(cells)

            # Find UUID in this row
            uuid = _extract_uuid(row_text)
            if not uuid:
                continue
            if uuid in uuid_map:
                continue  # already found from a higher-priority sheet

            # Try to get a name — heuristic: look for common name columns
            name = None
            dataset = sheet_name

            # Special handling for known sheet layouts
            if sheet_name == "Foundational Model":
                # cols: Dataset, Name, Video Link
                dataset = str(row[0]) if row[0] else sheet_name
                name = str(row[1]) if row[1] else None
            elif sheet_name == "Vestibular Schwannoma":
                # cols: Sortable Code, Code, ..., Name for download
                name = str(row[0]) if row[0] else None  # Sortable Code (RS-034)
                dataset = "VS"
            elif sheet_name == "MVD":
                # cols: Upload Taxonomy, Link, UUID, ...
                name = str(row[0]) if row[0] else None  # Upload Taxonomy
                dataset = "MVD"
            elif sheet_name == "5ALA HGG Resections":
                name = str(row[0]) if row[0] else None
                dataset = "Tumour_Resections"
            elif sheet_name == "ATLR":
                name = str(row[0]) if row[0] else None
                dataset = "ATLR"
            elif sheet_name == "Aneurysm Clipping (real)":
                # cols: ID, MRN, ..., Name
                name = str(row[12]) if len(row) > 12 and row[12] else None
                dataset = "Aneurysm_Clipping"
            else:
                # Generic: take first non-empty non-UUID non-link cell as name
                for c in row:
                    s = str(c).strip() if c else ""
                    if s and not _UUID_RE.search(s) and "touchsurgery" not in s and len(s) > 1:
                        name = s
                        break

            uuid_map[uuid] = {
                "name": name or f"unknown_{uuid[:8]}",
                "dataset": dataset,
                "sheet": sheet_name,
            }

    wb.close()
    return uuid_map


def load_phase_labels(phases_path: Path = PHASES_PATH) -> list[dict]:
    """Load phase-labels.json and return the sequences list."""
    with open(phases_path) as f:
        data = json.load(f)
    return data["sequences"]


def map_phases_to_videos(
    phases_path: Path = PHASES_PATH,
    excel_path: Path = EXCEL_PATH,
) -> list[dict]:
    """Map each phase-label sequence to its video name and metadata.

    Returns list of dicts with keys:
        video_id, name, dataset, sheet, phases (list of {label, start, end})
    """
    uuid_map = build_uuid_to_info(excel_path)
    sequences = load_phase_labels(phases_path)

    results = []
    found = 0
    missing = []

    for seq in sequences:
        vid = seq["videoId"].lower()
        info = uuid_map.get(vid)

        phases = []
        for ann in seq.get("videoAnnotations", []):
            phases.append({
                "label": ann["label"]["displayName"],
                "code": ann["label"]["code"],
                "start_ms": ann["timestampStart"],
                "end_ms": ann["timestampEnd"],
            })
        phases.sort(key=lambda p: p["start_ms"])

        if info:
            found += 1
            results.append({
                "video_id": vid,
                "name": info["name"],
                "dataset": info["dataset"],
                "sheet": info["sheet"],
                "phases": phases,
            })
        else:
            missing.append(vid)
            results.append({
                "video_id": vid,
                "name": None,
                "dataset": None,
                "sheet": None,
                "phases": phases,
            })

    return results, found, missing


# ---------------------------------------------------------------------------
# Video name → frame directory matching
# ---------------------------------------------------------------------------

IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".tiff"}


def _prefix_key(name: str) -> str:
    """Extract prefix before first underscore (or whole name)."""
    idx = name.find("_")
    return name[:idx] if idx > 0 else name


def match_video_to_dir(
    name: str,
    dataset: str | None,
    available_keys: list[str],
) -> str | None:
    """Match a phase-label video name to an extracted video directory key.

    Tries: exact match on dir name, then prefix match, then contains.
    """
    name_lower = name.lower()
    prefix = _prefix_key(name).lower()

    # Try exact match on last path component
    for key in available_keys:
        dir_name = key.rsplit("/", 1)[-1].lower()
        if dir_name == name_lower:
            return key

    # Prefix match (RS-034 matches RS-034_vestibular_schwannoma_...)
    for key in available_keys:
        dir_name = key.rsplit("/", 1)[-1].lower()
        if dir_name.startswith(prefix + "_") or dir_name.startswith(prefix + "-"):
            return key
        if name_lower in dir_name:
            return key

    return None


def discover_available_keys(
    frames_dir: Path,
    categories: list[str] | None = None,
) -> list[str]:
    """Cheap directory scan — returns video keys without reading any images."""
    keys = []
    for item in sorted(frames_dir.iterdir()):
        if not item.is_dir():
            continue
        has_images = any(f.suffix.lower() in IMAGE_SUFFIXES for f in item.iterdir() if f.is_file())
        if has_images:
            keys.append(item.name)
        else:
            if categories and item.name not in categories:
                continue
            for vdir in sorted(item.iterdir()):
                if vdir.is_dir():
                    keys.append(f"{item.name}/{vdir.name}")
    return keys


def match_phases_to_frames(
    phase_data: list[dict],
    frames_dir: Path,
    categories: list[str] | None = None,
) -> tuple[dict[str, str], dict[str, list[str]]]:
    """Match phase-labeled videos to frame directories.

    Returns:
        video_matches: {video_name: dir_key} for matched videos
        dir_key_users: {dir_key: [video_names]} showing collisions
    """
    available_keys = discover_available_keys(frames_dir, categories)

    video_matches: dict[str, str] = {}
    dir_key_users: dict[str, list[str]] = {}

    for vinfo in phase_data:
        if vinfo["name"] is None:
            continue
        dir_key = match_video_to_dir(vinfo["name"], vinfo.get("dataset"), available_keys)
        if dir_key is not None:
            video_matches[vinfo["name"]] = dir_key
            dir_key_users.setdefault(dir_key, []).append(vinfo["name"])

    return video_matches, dir_key_users


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Map phase labels to video names")
    parser.add_argument("--phases", type=Path, default=PHASES_PATH)
    parser.add_argument("--excel", type=Path, default=EXCEL_PATH)
    parser.add_argument("--frames-dir", type=Path, default=None,
                        help="Frames directory to match against")
    parser.add_argument("--categories", nargs="*", default=None)
    parser.add_argument("--output", type=Path, default=None,
                        help="Save mapped phases as JSON")
    args = parser.parse_args()

    results, found, missing = map_phases_to_videos(args.phases, args.excel)

    print(f"\n{'='*70}")
    print(f"Phase labels: {len(results)} sequences")
    print(f"Matched to Excel: {found}/{len(results)}")
    print(f"{'='*70}\n")

    # Match to frame directories if available
    frames_dir = args.frames_dir
    if frames_dir is None:
        from src.config import _find_existing, _FRAMES_CANDIDATES
        frames_dir = _find_existing(_FRAMES_CANDIDATES, None)

    video_matches: dict[str, str] = {}
    dir_key_users: dict[str, list[str]] = {}
    if frames_dir and frames_dir.is_dir():
        video_matches, dir_key_users = match_phases_to_frames(
            results, frames_dir, args.categories,
        )

    # Group by dataset
    by_dataset: dict[str, list] = {}
    for r in results:
        ds = r["dataset"] or "UNKNOWN"
        by_dataset.setdefault(ds, []).append(r)

    for ds in sorted(by_dataset):
        vids = by_dataset[ds]
        print(f"  {ds} ({len(vids)} videos):")
        for v in vids:
            n_phases = len(v["phases"])
            phase_labels = [p["label"] for p in v["phases"]]
            dir_key = video_matches.get(v["name"]) if v["name"] else None
            status = f"-> {dir_key}" if dir_key else "!! NO FRAMES"
            print(f"    {v['name'] or '(no name)':<55} {status}")
            print(f"      {n_phases} phases: {', '.join(phase_labels)}")

    # Warnings
    n_no_name = sum(1 for r in results if r["name"] is None)
    if n_no_name:
        print(f"\n  WARN: {n_no_name} videos have no name (UUID not found in Excel)")

    if missing:
        print(f"\n  MISSING ({len(missing)} UUIDs not found in Excel):")
        for vid in missing:
            print(f"    {vid}")

    for dk, names in dir_key_users.items():
        if len(names) > 1:
            print(f"\n  COLLISION: {len(names)} videos map to same dir {dk}:")
            for n in names:
                print(f"    - {n}")

    if video_matches:
        unique_dirs = set(video_matches.values())
        n_named = sum(1 for r in results if r["name"] is not None)
        n_miss = n_named - len(video_matches)
        print(f"\n  Summary: {len(results)} total, {n_named} named, "
              f"{len(video_matches)} matched to dirs ({len(unique_dirs)} unique), "
              f"{n_miss} missing frames")

    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        with open(args.output, "w") as f:
            json.dump(results, f, indent=2)
        print(f"\nSaved to {args.output}")


if __name__ == "__main__":
    main()
